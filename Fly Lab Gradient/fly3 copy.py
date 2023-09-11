import openpyxl
import math
import pandas as pd

#specifcy the path where original xlsx file exists
path = "/Users/navneet/Desktop/workbook2.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

#write 0 to every single value in the excel file
for row in ws.iter_rows(min_row=1, max_row=972, min_col=1, max_col=1296):
    for cell in row:
        cell.value=0

#change the decay formula to one desired               
#y = [(0.0003*x**2-0.348*x+100) for x in range(500)]
#y = [(-16*math.log(x+1)+100) for x in range(500)]
#y = [(100*math.exp(-0.008*x)) for x in range(500)]
#y = [(0.0004*x**2-0.4*x+100) for x in range(500)]

#applies the decay formula to the center of the sheet
for j in range(500):
        i = j+398
        for row2 in ws.iter_rows(min_row=336, max_row=636, min_col=i, max_col=i):
            for cell3 in row2:
                cell3.value=list(y)[j]

#saves the xlsx file
wb.save("workbook2.xlsx")

#converts the xlsx file to a csv
read_file = pd.read_excel (path)
read_file.to_csv ('exponential.csv', index = None, header=False)

#plot in 3d 
#change the equation x^-2 and inverse quadratic equationx